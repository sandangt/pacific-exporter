from typing import List

from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from app.model import LearningResult, Student
from app.repository import StudentRepository, LearningResultRepository
from app.utils import make_student_slug, map_grade, parse_mark, parse_teacher, parse_subject


class PersistService:
    def __init__(self, student_repository: StudentRepository, learning_result_repository: LearningResultRepository):
        self._student_repository = student_repository
        self._learning_result_repository = learning_result_repository

    @staticmethod
    def extract_sheet_student(sheet: Worksheet) -> List[Student]:
        class_code = sheet.title
        semester_title = sheet.cell(row=1, column=column_index_from_string('A')).value
        student_list = []
        current = 4
        while no_in_class := sheet.cell(row=current, column=column_index_from_string('A')).value:
            student_name = sheet.cell(current, column=column_index_from_string('B')).value
            student_list.append(Student(
                class_code=class_code,
                semester_title=semester_title,
                slug=make_student_slug(class_code, student_name),
                name=student_name,
                no_in_class=no_in_class,
            ))
            current += 1
        return student_list

    def extract_sheet_learning_result(self, sheet: Worksheet) -> List[LearningResult]:
        teacher_name = parse_teacher(sheet.cell(row=2, column=column_index_from_string('C')).value)
        subject_name = parse_subject(sheet.cell(row=2, column=column_index_from_string('A')).value)
        class_code = sheet.title
        learning_result_list = []
        current = 4
        while sheet.cell(row=current, column=column_index_from_string('A')).value:
            student_name = sheet.cell(current, column=column_index_from_string('B')).value
            student_slug = make_student_slug(class_code, student_name)
            student_entity = self._student_repository.get_one_by_slug(student_slug)
            mark = parse_mark(sheet.cell(row=current, column=column_index_from_string('D')).value)
            learning_result_list.append(LearningResult(
                mark=mark,
                grade=map_grade(mark),
                teacher_name=teacher_name,
                subject=subject_name,
                comment=sheet.cell(row=current, column=column_index_from_string('E')).value,
                student_id=student_entity.id
            ))
            current += 1
        return learning_result_list

    def import_sheet(self, sheet: Worksheet):
        student_list = self.extract_sheet_student(sheet)
        self._student_repository.upsert_multiple(student_list)
        learning_result_list = self.extract_sheet_learning_result(sheet)
        self._learning_result_repository.create_multiple(learning_result_list)

    def import_workbook(self, file_path: str):
        wb = load_workbook(file_path, data_only=True)
        for sheet in wb.worksheets:
            self.import_sheet(sheet)
