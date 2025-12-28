from datetime import datetime
from typing import List, Any

from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from slugify import slugify

from app.constant import SUBJECT_RANKING_MAP, MAX_INTEGER
from app.exception import InvalidMarkException
from app.model import LearningResult, Student
from app.repository import StudentRepository, LearningResultRepository
from app.utils import make_student_slug


class PersistService:
    def __init__(self, student_repository: StudentRepository, learning_result_repository: LearningResultRepository):
        self.__student_repository = student_repository
        self.__learning_result_repository = learning_result_repository

    def import_workbook(self, file_path: str):
        wb = load_workbook(file_path, data_only=True)
        for sheet in wb.worksheets:
            self.__import_sheet(sheet)

    @staticmethod
    def __extract_sheet_student(sheet: Worksheet) -> List[Student]:
        class_code = sheet.title
        semester_title = sheet.cell(row=1, column=column_index_from_string('A')).value
        student_list = []
        current = 4
        while no_in_class := sheet.cell(row=current, column=column_index_from_string('A')).value:
            student_name = sheet.cell(current, column=column_index_from_string('B')).value
            student_list.append(Student(
                class_code=class_code,
                semester_title=semester_title,
                slug=make_student_slug(class_code, student_name),
                name=student_name,
                no_in_class=no_in_class,
            ))
            current += 1
        return student_list

    def __extract_sheet_learning_result(self, sheet: Worksheet) -> List[LearningResult]:
        teacher_name = self.__parse_teacher(sheet.cell(row=2, column=column_index_from_string('C')).value)
        subject_name = self.__parse_subject(sheet.cell(row=2, column=column_index_from_string('A')).value)
        subject_slug = slugify(subject_name)
        subject_rank = SUBJECT_RANKING_MAP.get(subject_slug, MAX_INTEGER)
        class_code = sheet.title
        learning_result_list = []
        current = 4
        while sheet.cell(row=current, column=column_index_from_string('A')).value:
            student_name = sheet.cell(current, column=column_index_from_string('B')).value
            student_slug = make_student_slug(class_code, student_name)
            student_entity = self.__student_repository.get_one_by_slug(student_slug)
            mark = self.__parse_mark(sheet.cell(row=current, column=column_index_from_string('D')).value)
            learning_result_list.append(LearningResult(
                mark=mark,
                grade=self.__map_grade(mark),
                teacher_name=teacher_name,
                subject=subject_name,
                comment=sheet.cell(row=current, column=column_index_from_string('E')).value,
                student_id=student_entity.id,
                subject_slug=subject_slug,
                subject_rank=subject_rank,
            ))
            current += 1
        return learning_result_list

    def __import_sheet(self, sheet: Worksheet):
        student_list = self.__extract_sheet_student(sheet)
        self.__student_repository.upsert_multiple(student_list)
        learning_result_list = self.__extract_sheet_learning_result(sheet)
        self.__learning_result_repository.create_multiple(learning_result_list)

    @staticmethod
    def __map_grade(mark: float) -> str:
        match mark:
            case val if 9 <= val <= 10:
                return 'A'
            case val if 8 <= val <= 8.9:
                return 'B'
            case val if 7 <= val <= 7.9:
                return 'C'
            case val if 5 <= val <= 6.9:
                return 'D'
            case val if 0 <= val <= 4.9:
                return 'F'
            case _:
                raise InvalidMarkException.default()

    @staticmethod
    def __parse_mark(mark: Any) -> float:
        if isinstance(mark, datetime):
            return float(str(mark.day) + '.' + str(mark.month))
        return float(mark)

    @staticmethod
    def __parse_subject(cell: str) -> str:
        return cell.split(':')[1].strip()

    @staticmethod
    def __parse_teacher(cell: str) -> str:
        return cell.split(':')[1].strip()
